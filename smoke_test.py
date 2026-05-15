"""Manual smoke test: run extractor on all fixtures and print table counts + first row preview."""
from pathlib import Path
from pdf2xlsx.extractor import extract_tables
import openpyxl
from pdf2xlsx.writer import write_xlsx
import tempfile

FIXTURES = Path("tests/fixtures")

for pdf in sorted(FIXTURES.glob("*.pdf")):
    print(f"\n{'='*60}")
    print(f"PDF: {pdf.name} ({pdf.stat().st_size:,} bytes)")
    tables = extract_tables(pdf)
    print(f"Tables found: {len(tables)}")
    for t in tables[:5]:  # show first 5
        first_row = " | ".join(str(c)[:20] for c in t.rows[0][:4])
        print(f"  p.{t.page} [{t.source}] {t.row_count}r×{t.col_count}c: {first_row}")
    if len(tables) > 5:
        print(f"  ... and {len(tables)-5} more tables")
    # Write and verify XLSX
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        out = Path(f.name)
    write_xlsx(tables, out)
    wb = openpyxl.load_workbook(out)
    print(f"XLSX: {len(wb.sheetnames)} sheets ({', '.join(wb.sheetnames[:4])}{'...' if len(wb.sheetnames)>4 else ''})")
    out.unlink()

print("\nAll PDFs processed successfully.")
