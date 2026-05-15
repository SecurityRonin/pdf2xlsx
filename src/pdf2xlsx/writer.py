from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pdf2xlsx.models import ExtractedTable

_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_INVALID_CHARS = str.maketrans({"/": "-", "\\": "-", "?": "", "*": "", "[": "", "]": "", ":": ""})


def _safe_sheet_name(name: str) -> str:
    return name.translate(_INVALID_CHARS)[:31]


def _write_table_sheet(ws, table: ExtractedTable) -> None:
    for row_idx, row in enumerate(table.rows, start=1):
        for col_idx, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(cell_value))
            if row_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
                cell.alignment = Alignment(wrap_text=True)


def _write_index_sheet(ws, tables: list[ExtractedTable]) -> None:
    headers = ["Sheet", "Page", "Rows", "Columns", "Source"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL

    for row_idx, table in enumerate(tables, start=2):
        ws.cell(row=row_idx, column=1, value=table.sheet_name)
        ws.cell(row=row_idx, column=2, value=table.page)
        ws.cell(row=row_idx, column=3, value=table.row_count)
        ws.cell(row=row_idx, column=4, value=table.col_count)
        ws.cell(row=row_idx, column=5, value=table.source)


def write_xlsx(tables: list[ExtractedTable], output_path: Path) -> None:
    output_path = Path(output_path)
    wb = openpyxl.Workbook()

    ws_index = wb.active
    ws_index.title = "Index"
    _write_index_sheet(ws_index, tables)

    used_names: dict[str, int] = {}
    for table in tables:
        base_name = _safe_sheet_name(table.sheet_name)
        if base_name in used_names:
            used_names[base_name] += 1
            suffix = f"-{used_names[base_name]}"
            name = _safe_sheet_name(base_name[:31 - len(suffix)] + suffix)
        else:
            used_names[base_name] = 1
            name = base_name

        ws = wb.create_sheet(title=name)
        _write_table_sheet(ws, table)

    wb.save(output_path)
