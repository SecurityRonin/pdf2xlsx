"""Integration tests against all downloaded real PDFs."""
import pytest
from pathlib import Path
import openpyxl
from pdf2xlsx.extractor import extract_tables
from pdf2xlsx.writer import write_xlsx

FIXTURES = Path(__file__).parent / "fixtures"
ALL_PDFS = [
    "annual_report.pdf",
    "esg_report.pdf",
    "academic_paper.pdf",
    "esg_disclosure.pdf",
    "term_sheet.pdf",
    "general_ledger.pdf",
]


@pytest.mark.parametrize("pdf_name", ALL_PDFS)
def test_pdf_produces_valid_xlsx(pdf_name, tmp_path):
    pdf_path = FIXTURES / pdf_name
    if not pdf_path.exists():
        pytest.skip(f"Fixture not downloaded: {pdf_name}")

    tables = extract_tables(pdf_path)
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)

    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert "Index" in wb.sheetnames


@pytest.mark.parametrize("pdf_name", ALL_PDFS)
def test_pdf_has_at_least_one_table(pdf_name):
    pdf_path = FIXTURES / pdf_name
    if not pdf_path.exists():
        pytest.skip(f"Fixture not downloaded: {pdf_name}")

    tables = extract_tables(pdf_path)
    assert len(tables) > 0, f"{pdf_name}: expected tables, got 0"


@pytest.mark.parametrize("pdf_name", ALL_PDFS)
def test_pdf_tables_all_have_data(pdf_name):
    pdf_path = FIXTURES / pdf_name
    if not pdf_path.exists():
        pytest.skip(f"Fixture not downloaded: {pdf_name}")

    tables = extract_tables(pdf_path)
    for t in tables:
        assert t.row_count >= 2, (
            f"{pdf_name} p.{t.page}: table has only {t.row_count} row(s)"
        )
        assert t.col_count >= 1, (
            f"{pdf_name} p.{t.page}: table has {t.col_count} columns"
        )


@pytest.mark.parametrize("pdf_name", ALL_PDFS)
def test_pdf_no_none_cells(pdf_name):
    pdf_path = FIXTURES / pdf_name
    if not pdf_path.exists():
        pytest.skip(f"Fixture not downloaded: {pdf_name}")

    tables = extract_tables(pdf_path)
    for t in tables:
        for row in t.rows:
            for cell in row:
                assert cell is not None, (
                    f"{pdf_name} p.{t.page}: None cell found"
                )


@pytest.mark.parametrize("pdf_name", ALL_PDFS)
def test_xlsx_sheet_names_all_valid(pdf_name, tmp_path):
    pdf_path = FIXTURES / pdf_name
    if not pdf_path.exists():
        pytest.skip(f"Fixture not downloaded: {pdf_name}")

    tables = extract_tables(pdf_path)
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)
    wb = openpyxl.load_workbook(out)
    for name in wb.sheetnames:
        assert len(name) <= 31
        assert "/" not in name
        assert "\\" not in name
        assert "?" not in name
        assert "*" not in name
        assert "[" not in name
        assert "]" not in name
