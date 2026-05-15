import pytest
from typer.testing import CliRunner
from pdf2xlsx.cli import app
from pathlib import Path
import openpyxl
import shutil

runner = CliRunner()
FIXTURES = Path(__file__).parent / "fixtures"


def test_cli_help():
    result = runner.invoke(app, ["--help"])
    assert result.exit_code == 0
    assert "pdf" in result.output.lower() or "xlsx" in result.output.lower()


def test_cli_missing_input_file(tmp_path):
    result = runner.invoke(app, [str(tmp_path / "nonexistent.pdf"), str(tmp_path / "out.xlsx")])
    assert result.exit_code != 0


def test_cli_produces_xlsx(tmp_path):
    out = tmp_path / "output.xlsx"
    result = runner.invoke(app, [str(FIXTURES / "academic_paper.pdf"), str(out)])
    assert result.exit_code == 0, result.output
    assert out.exists()


def test_cli_output_has_index_sheet(tmp_path):
    out = tmp_path / "output.xlsx"
    runner.invoke(app, [str(FIXTURES / "annual_report.pdf"), str(out)])
    wb = openpyxl.load_workbook(out)
    assert "Index" in wb.sheetnames


def test_cli_output_has_multiple_sheets(tmp_path):
    out = tmp_path / "output.xlsx"
    runner.invoke(app, [str(FIXTURES / "annual_report.pdf"), str(out)])
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) >= 2


def test_cli_default_output_path(tmp_path):
    src = tmp_path / "academic_paper.pdf"
    shutil.copy(FIXTURES / "academic_paper.pdf", src)
    result = runner.invoke(app, [str(src)])
    assert result.exit_code == 0, result.output
    expected_out = tmp_path / "academic_paper.xlsx"
    assert expected_out.exists()
