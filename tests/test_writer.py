import pytest
from pathlib import Path
import openpyxl
from pdf2xlsx.writer import write_xlsx
from pdf2xlsx.models import ExtractedTable


@pytest.fixture
def two_tables():
    return [
        ExtractedTable(
            page=1, index=0,
            rows=[["Name", "Q1", "Q2"], ["Revenue", "100", "120"], ["Costs", "80", "90"]],
            source="pdfplumber",
        ),
        ExtractedTable(
            page=2, index=0,
            rows=[["Category", "Score"], ["Env", "A"], ["Social", "B"]],
            source="pdfplumber",
        ),
    ]


def test_write_creates_file(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    assert out.exists()


def test_write_creates_index_sheet(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    assert "Index" in wb.sheetnames


def test_write_index_is_first_sheet(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames[0] == "Index"


def test_write_creates_one_sheet_per_table(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    assert len(wb.sheetnames) == 3  # Index + 2 tables


def test_write_sheet_names_match_tables(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    assert "Table 1 (p.1)" in wb.sheetnames
    assert "Table 1 (p.2)" in wb.sheetnames


def test_write_table_data_correct(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Table 1 (p.1)"]
    assert ws.cell(1, 1).value == "Name"
    assert ws.cell(2, 1).value == "Revenue"
    assert ws.cell(3, 3).value == "90"


def test_write_index_has_headers(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Index"]
    assert ws.cell(1, 1).value == "Sheet"


def test_write_index_lists_all_tables(tmp_path, two_tables):
    out = tmp_path / "out.xlsx"
    write_xlsx(two_tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Index"]
    assert ws.cell(2, 1).value == "Table 1 (p.1)"
    assert ws.cell(3, 1).value == "Table 1 (p.2)"


def test_write_empty_tables_list(tmp_path):
    out = tmp_path / "out.xlsx"
    write_xlsx([], out)
    wb = openpyxl.load_workbook(out)
    assert "Index" in wb.sheetnames
    assert len(wb.sheetnames) == 1


def test_write_numeric_strings_preserved(tmp_path):
    tables = [ExtractedTable(
        page=1, index=0,
        rows=[["Item", "Amount"], ["Tax", "1,234.56"]],
        source="pdfplumber",
    )]
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Table 1 (p.1)"]
    assert ws.cell(2, 2).value == "1,234.56"


def test_write_duplicate_sheet_names_disambiguated(tmp_path):
    tables = [
        ExtractedTable(page=1, index=0, rows=[["A"], ["1"]], source="pdfplumber"),
        ExtractedTable(page=1, index=0, rows=[["B"], ["2"]], source="pymupdf"),
    ]
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)
    wb = openpyxl.load_workbook(out)
    # Index + 2 distinct sheets
    assert len(wb.sheetnames) == 3
    assert len(set(wb.sheetnames)) == 3  # all unique


def test_write_sheet_names_under_31_chars(tmp_path):
    tables = [ExtractedTable(page=999, index=99, rows=[["H"], ["V"]], source="pdfplumber")]
    out = tmp_path / "out.xlsx"
    write_xlsx(tables, out)
    wb = openpyxl.load_workbook(out)
    for name in wb.sheetnames:
        assert len(name) <= 31, f"Sheet name too long: {name!r}"
